"""
Generador de Reportes Excel para el Analizador BBPP
Incluye hojas de resumen, hallazgos y estadísticas con gráficos
"""

from pathlib import Path
from typing import Dict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelReportGenerator:
    """Generador de reportes Excel para análisis de BBPP"""
    
    def __init__(self, results: Dict, output_path: Path = None, include_charts: bool = True):
        """
        Inicializar generador
        
        Args:
            results: Resultados del análisis
            output_path: Ruta donde guardar el reporte (opcional)
            include_charts: Si incluir gráficos
        """
        self.results = results
        self.include_charts = include_charts and OPENPYXL_AVAILABLE
        
        # Cargar colores desde branding
        try:
            from src.branding_manager import get_color
            self.COLOR_PRIMARY = get_color('primary').replace('#', '')
            self.COLOR_SECONDARY = get_color('secondary').replace('#', '')
        except ImportError:
            # Fallback colors
            self.COLOR_PRIMARY = "00215F"
            self.COLOR_SECONDARY = "0077B6"
        
        # Colores para severidades (fijos)
        self.COLOR_SUCCESS = "28A745"  # Verde
        self.COLOR_WARNING = "FFC107"  # Amarillo
        self.COLOR_ERROR = "DC3545"    # Rojo
        self.COLOR_INFO = "0D6EFD"     # Azul (cambiado de cyan a azul)
        self.COLOR_HEADER = "E8E8E8"   # Gris claro
        
        # Si no se especifica ruta, usar estructura nueva con nombre estandarizado
        if output_path is None:
            from src.report_utils import get_report_output_dir, generate_report_filename
            project_name = results.get('project_info', {}).get('name', 'Proyecto')
            output_dir = get_report_output_dir('excel')
            filename = generate_report_filename(project_name, 'xlsx')
            self.output_path = output_dir / filename
        else:
            self.output_path = output_path
        
        # Estilos predefinidos (se configuran aquí, pero el workbook se crea en generate)
        self._setup_styles()
    
    def _setup_styles(self):
        """Configurar estilos reutilizables"""
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color=self.COLOR_PRIMARY, 
                                        end_color=self.COLOR_PRIMARY, 
                                        fill_type="solid")
        self.subheader_fill = PatternFill(start_color=self.COLOR_HEADER,
                                           end_color=self.COLOR_HEADER,
                                           fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
    
    def generate(self) -> Path:
        """
        Generar el reporte Excel
        
        Returns:
            Path al archivo generado
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Instala con: pip install openpyxl")
        
        # Crear workbook
        self.wb = Workbook()
        
        # Asegurar que existe la carpeta output
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Crear hojas
        self._create_summary_sheet()
        self._create_findings_sheet()
        self._create_statistics_sheet()
        self._create_files_sheet()
        
        # Eliminar hoja por defecto si existe
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        
        # Guardar archivo
        self.wb.save(self.output_path)
        
        return self.output_path
    
    def _create_summary_sheet(self):
        """Crear hoja de resumen ejecutivo"""
        ws = self.wb.create_sheet("Resumen", 0)
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = "📊 RESUMEN EJECUTIVO - ANÁLISIS DE BUENAS PRÁCTICAS"
        ws['A1'].font = Font(bold=True, size=16, color=self.COLOR_PRIMARY)
        ws['A1'].alignment = self.center_align
        ws.row_dimensions[1].height = 30
        
        # Información del proyecto
        row = 3
        project_info = self.results.get('project_info', {})
        score_data = self.results.get('score', {})
        stats = self.results.get('statistics', {})
        
        # Sección: Información del Proyecto
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "📁 INFORMACIÓN DEL PROYECTO"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        info_rows = [
            ("Nombre del Proyecto:", project_info.get('name', 'N/A')),
            ("Tipo:", project_info.get('type', 'N/A')),
            ("UiPath Studio:", project_info.get('studio_version', 'N/A')),
            ("Archivos Analizados:", self.results.get('analyzed_files', 0)),
            ("Fecha de Análisis:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]
        
        for label, value in info_rows:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Sección: Score
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "🎯 PUNTUACIÓN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        score = score_data.get('score', 0)
        grade = score_data.get('grade', 'N/A')
        
        ws[f'A{row}'] = "Score:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{score}/100"
        ws[f'B{row}'].font = Font(bold=True, size=14)
        
        # Color según score
        if score >= 90:
            ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLOR_SUCCESS)
        elif score >= 70:
            ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLOR_WARNING)
        else:
            ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLOR_ERROR)
        row += 1
        
        ws[f'A{row}'] = "Calificación:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = grade
        row += 2
        
        # Sección: Hallazgos
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "⚠️ RESUMEN DE HALLAZGOS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        findings_rows = [
            ("❌ Errores:", stats.get('errors', 0), self.COLOR_ERROR, "FFE6E6"),
            ("⚠️ Warnings:", stats.get('warnings', 0), "B8860B", "FFF9E6"),
            ("ℹ️ Info:", stats.get('infos', 0), self.COLOR_INFO, "E6F2FF"),
            ("📊 Total:", stats.get('total_findings', 0), self.COLOR_PRIMARY, "E8E8E8"),
        ]
        
        for label, value, color, bg_color in findings_rows:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = self.border
            ws[f'A{row}'].alignment = self.left_align
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(bold=True, color=color)
            ws[f'B{row}'].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            ws[f'B{row}'].border = self.border
            ws[f'B{row}'].alignment = self.center_align
            row += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        
        # Gráfico de severidades si hay hallazgos y gráficos habilitados
        if self.include_charts and stats.get('total_findings', 0) > 0:
            self._add_severity_chart(ws, stats, start_row=3, start_col=4)
    
    def _add_severity_chart(self, ws, stats, start_row, start_col):
        """Añadir gráfico de severidades"""
        # Datos para el gráfico
        data_start_row = start_row
        ws.cell(row=data_start_row, column=start_col, value="Severidad")
        ws.cell(row=data_start_row, column=start_col + 1, value="Cantidad")
        
        data_rows = [
            ("Errores", stats.get('errors', 0)),
            ("Warnings", stats.get('warnings', 0)),
            ("Info", stats.get('infos', 0)),
        ]
        
        for i, (label, value) in enumerate(data_rows, 1):
            ws.cell(row=data_start_row + i, column=start_col, value=label)
            ws.cell(row=data_start_row + i, column=start_col + 1, value=value)
        
        # Crear gráfico de pastel
        chart = PieChart()
        chart.title = "Distribución por Severidad"
        
        data = Reference(ws, min_col=start_col + 1, min_row=data_start_row, 
                        max_row=data_start_row + len(data_rows))
        cats = Reference(ws, min_col=start_col, min_row=data_start_row + 1, 
                        max_row=data_start_row + len(data_rows))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = True
        
        # Aplicar colores personalizados usando el método de series
        # Este método es más compatible con OnlyOffice y otros lectores
        from openpyxl.chart.series import DataPoint
        
        # Colores: Rojo para Errores, Amarillo para Warnings, Azul para Info
        colors = [
            self.COLOR_ERROR,    # Rojo para Errores
            "FFC107",            # Amarillo para Warnings
            self.COLOR_INFO      # Azul para Info
        ]
        
        # Crear puntos de datos con colores específicos
        try:
            # Método alternativo más compatible
            for idx, color in enumerate(colors):
                pt = DataPoint(idx=idx)
                pt.graphicalProperties = GraphicalProperties()
                pt.graphicalProperties.solidFill = color
                chart.series[0].dPt.append(pt)
        except Exception as e:
            print(f"Warning: Método 1 falló, intentando método 2: {e}")
            # Método alternativo 2: Usar el estilo de la serie
            try:
                chart.series[0].graphicalProperties = GraphicalProperties()
                chart.series[0].graphicalProperties.solidFill = self.COLOR_ERROR
            except Exception as e2:
                print(f"Warning: No se pudieron aplicar colores al gráfico: {e2}")
        
        # Posición del gráfico
        ws.add_chart(chart, f"{get_column_letter(start_col + 3)}{start_row}")
    
    def _create_findings_sheet(self):
        """Crear hoja con todos los hallazgos"""
        ws = self.wb.create_sheet("Hallazgos")
        
        # Headers
        headers = ["#", "Severidad", "Categoría", "Descripción", "Archivo", "Ubicación"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        # Datos
        findings = self.results.get('findings', [])
        
        # Color de fila alternada
        alternate_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_idx, finding in enumerate(findings, 2):
            severity = finding.get('severity', 'info')
            
            # Icono de severidad
            severity_icon = {
                'error': '❌ Error',
                'warning': '⚠️ Warning',
                'info': 'ℹ️ Info'
            }.get(severity, severity)
            
            row_data = [
                row_idx - 1,
                severity_icon,
                finding.get('category', ''),
                finding.get('description', ''),
                Path(finding.get('file_path', '')).name,
                finding.get('location', '')
            ]
            
            # Determinar color de fondo para la celda de severidad
            if severity == 'error':
                severity_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                severity_font = Font(color=self.COLOR_ERROR, bold=True)
            elif severity == 'warning':
                severity_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                severity_font = Font(color="B8860B", bold=True)  # Dorado oscuro para mejor contraste
            else:  # info
                severity_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
                severity_font = Font(color=self.COLOR_INFO, bold=True)
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = self.left_align if col_idx > 2 else self.center_align
                
                # Aplicar color de fondo alternado
                if row_idx % 2 == 0:
                    cell.fill = alternate_fill
                
                # Color especial para columna de severidad
                if col_idx == 2:
                    cell.fill = severity_fill
                    cell.font = severity_font
                    cell.alignment = self.center_align
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        
        # Filtros automáticos
        ws.auto_filter.ref = f"A1:F{len(findings) + 1}"
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
    
    def _create_statistics_sheet(self):
        """Crear hoja de estadísticas"""
        ws = self.wb.create_sheet("Estadísticas")
        
        stats = self.results.get('statistics', {})
        
        # Título
        ws.merge_cells('A1:C1')
        ws['A1'] = "📈 ESTADÍSTICAS DETALLADAS"
        ws['A1'].font = Font(bold=True, size=14, color=self.COLOR_PRIMARY)
        
        row = 3
        
        # Estadísticas generales
        ws[f'A{row}'] = "MÉTRICAS GENERALES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        general_stats = [
            ("Total Actividades:", stats.get('total_activities', 0)),
            ("Total Variables:", stats.get('total_variables', 0)),
            ("Total Argumentos:", stats.get('total_arguments', 0)),
            ("Bloques Try-Catch:", stats.get('total_try_catch', 0)),
            ("Total LogMessages:", stats.get('total_logs', 0)),
            ("Archivos con Código Comentado:", stats.get('files_with_commented_code', 0)),
            ("Total Líneas Comentadas:", stats.get('total_commented_lines', 0)),
        ]
        
        for label, value in general_stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Hallazgos por categoría
        ws[f'A{row}'] = "HALLAZGOS POR CATEGORÍA"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        by_category = stats.get('by_category', {})
        chart_data_start = row
        for category, count in by_category.items():
            ws[f'A{row}'] = category.replace('_', ' ').title()
            ws[f'B{row}'] = count
            row += 1
        chart_data_end = row - 1
        
        # Gráfico de barras por categoría
        if self.include_charts and by_category:
            chart = BarChart()
            chart.title = "Hallazgos por Categoría"
            chart.type = "col"
            chart.style = 10
            
            data = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)
            cats = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
            
            chart.add_data(data)
            chart.set_categories(cats)
            chart.legend = None
            
            ws.add_chart(chart, "D3")
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
    
    def _create_files_sheet(self):
        """Crear hoja con información de archivos analizados"""
        ws = self.wb.create_sheet("Archivos")
        
        # Headers
        headers = ["Archivo", "Tipo", "Actividades", "Variables", "Argumentos", 
                   "Logs", "% Comentado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        # Datos de archivos parseados
        parsed_files = self.results.get('parsed_files', [])
        for row_idx, file_data in enumerate(parsed_files, 2):
            # Calcular % comentado
            total_lines = file_data.get('total_lines', 0)
            commented_lines = file_data.get('commented_lines', 0)
            comment_percent = (commented_lines / total_lines * 100) if total_lines > 0 else 0
            
            row_data = [
                Path(file_data.get('file_path', '')).name,
                file_data.get('workflow_type', 'Unknown'),
                len(file_data.get('activities', [])),
                len(file_data.get('variables', [])),
                len(file_data.get('arguments', [])),
                len(file_data.get('log_messages', [])),
                f"{comment_percent:.1f}%"
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                
                # Resaltar código comentado excesivo
                if col_idx == 7 and comment_percent > 5:
                    cell.font = Font(color=self.COLOR_WARNING, bold=True)
        
        # Ajustar anchos
        widths = [30, 15, 12, 12, 12, 10, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Filtros automáticos
        ws.auto_filter.ref = f"A1:G{len(parsed_files) + 1}"

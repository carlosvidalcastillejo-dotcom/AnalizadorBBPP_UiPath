    def _add_severity_chart(self, ws, stats, start_row, start_col):
        """Añadir gráfico de severidades con colores personalizados"""
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice
        
        # Datos para el gráfico
        data_start_row = start_row
        ws.cell(row=data_start_row, column=start_col, value="Severidad")
        ws.cell(row=data_start_row, column=start_col + 1, value="Cantidad")
        
        data_rows = [
            ("Errores", stats.get('errors', 0)),
            ("Warnings", stats.get('warnings', 0)),
            ("Info", stats.get('infos', 0)),
        ]
        
        for i, (label, value) in enumerate(data_rows, 1):
            ws.cell(row=data_start_row + i, column=start_col, value=label)
            ws.cell(row=data_start_row + i, column=start_col + 1, value=value)
        
        # Crear gráfico de pastel
        chart = PieChart()
        chart.title = "Distribución por Severidad"
        
        data = Reference(ws, min_col=start_col + 1, min_row=data_start_row, 
                        max_row=data_start_row + len(data_rows))
        cats = Reference(ws, min_col=start_col, min_row=data_start_row + 1, 
                        max_row=data_start_row + len(data_rows))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = True
        
        # Aplicar colores personalizados a cada serie
        # Colores: Rojo para Errores, Amarillo para Warnings, Azul para Info
        colors = [
            self.COLOR_ERROR,    # Rojo para Errores
            "FFC107",            # Amarillo para Warnings  
            self.COLOR_INFO      # Azul para Info
        ]
        
        # Aplicar colores a cada punto de datos
        try:
            for i, color in enumerate(colors):
                if i < len(chart.series[0].data_points):
                    point = chart.series[0].data_points[i]
                    # Crear propiedades gráficas con color sólido
                    fill = SolidColorFillProperties()
                    fill.solidFill = ColorChoice(srgbClr=color)
                    point.graphicalProperties = GraphicalProperties(solidFill=fill.solidFill)
        except Exception as e:
            # Si falla la asignación de colores, continuar sin colores personalizados
            print(f"Warning: No se pudieron aplicar colores personalizados al gráfico: {e}")
        
        # Posición del gráfico
        ws.add_chart(chart, f"{get_column_letter(start_col + 3)}{start_row}")

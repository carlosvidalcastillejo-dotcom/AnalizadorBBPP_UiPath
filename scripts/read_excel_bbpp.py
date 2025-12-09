"""
Script temporal para leer el archivo Excel con BBPP
"""
import openpyxl
from pathlib import Path
import sys
import json

# Configurar encoding UTF-8 para la salida
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

excel_path = Path(r"c:\Users\Imrik\Downloads\BBPP_UiPath_Implementacion (2).xlsx")

if not excel_path.exists():
    print(f"ERROR: No se encuentra el archivo {excel_path}")
    exit(1)

# Cargar workbook
wb = openpyxl.load_workbook(excel_path)

print(f"Hojas disponibles: {wb.sheetnames}\n")

# Leer cada hoja
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"HOJA: {sheet_name}")
    print(f"{'='*80}\n")

    # Obtener encabezados (primera fila)
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value is not None else '')

    print(f"Columnas: {headers}\n")
    print(f"Total filas: {ws.max_row}\n")

    # Leer todas las filas como diccionarios
    for row_idx in range(2, ws.max_row + 1):  # Empezar desde fila 2 (después de encabezados)
        row_data = {}
        has_data = False

        for col_idx, header in enumerate(headers, 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data[header] = str(cell_value).strip()
                has_data = True
            else:
                row_data[header] = ''

        # Solo mostrar filas con datos
        if has_data and row_data.get('ID', ''):
            print(f"\n--- Fila {row_idx} ---")
            for key, value in row_data.items():
                if value:
                    # Limitar longitud para mejor lectura
                    display_value = value if len(value) <= 200 else value[:200] + "..."
                    print(f"  {key}: {display_value}")

wb.close()
print("\n\nLectura completada.")
